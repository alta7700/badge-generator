import argparse
from pathlib import Path
from typing import Iterable

import openpyxl
import photoshop.api as ps


def set_visible(lines: Iterable[int]):
    all_lines = set(lines_layers.keys())
    for line in lines:
        lines_layers[line].visible = True
        all_lines.remove(line)
    for line in all_lines:
        lines_layers[line].visible = False


def fill_fio_layers(fio_lines: list[str]):
    match len(fio_lines):
        case 1:
            lines = (32,)
        case 2:
            lines = (21, 22)
        case 3:
            lines = (31, 32, 33)
        case 4:
            lines = (41, 42, 43, 44)
        case _:
            raise
    set_visible(lines)
    for i, line in enumerate(lines):
        lines_layers[line].textItem.contents = fio_lines[i]


def main(lines: list[int], results_dir: Path, path: Path, id_mode: bool):
    results_dir = results_dir.absolute()
    gen = iter_xlsx if path.name.endswith(".xlsx") else iter_csv
    for user_id, course, group, fio_lines in gen(path, id_mode):
        if id_mode:
            out_filename = f'{user_id}.png'
        else:
            out_filename = f'{course}к_{str(group).zfill(3)}гр_{" ".join(fio_lines)}.png'
        if len(fio_lines) not in lines:
            continue
        for layer in doc.artLayers:
            if layer.kind == ps.LayerKind.TextLayer:
                match layer.name:
                    case '__КУРС':
                        layer.textItem.contents = course
                    case '__ГРУППА':
                        layer.textItem.contents = group
        fill_fio_layers(fio_lines)
        png_file = results_dir / out_filename
        doc.saveAs(png_file.as_posix(), ps.PNGSaveOptions(), asCopy=True)
        print(png_file.name)


def iter_xlsx(path: Path, id_mode: bool) -> Iterable[tuple[str | None, str, str, list[str]]]:
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for i in range(2, sheet.max_row + 1):
        if id_mode:
            user_id = str(sheet.cell(row=i, column=1).value)
            start_col = 2
        else:
            user_id = None
            start_col = 1
        course = str(sheet.cell(i, start_col).value)
        group = str(sheet.cell(i, start_col + 1).value)
        fio_lines = [v for c in range(2, 6) if (v := sheet.cell(i, start_col + c).value)]
        yield user_id, course, group, fio_lines


def iter_csv(path: Path, id_mode: bool) -> Iterable[tuple[str | None, str, str, list[str]]]:
    with open(path, 'r', encoding='utf-8') as studs:
        studs.readline()  # read header
        for stud in studs:
            if not stud:
                continue
            if id_mode:
                user_id, course, group, f1, f2, f3, f4 = (x for x in stud.strip('\n').split(';'))
            else:
                user_id = None
                course, group, f1, f2, f3, f4 = (x for x in stud.strip('\n').split(';'))
            yield user_id, course, group, [v for v in (f1, f2, f3, f4) if v]


if __name__ == '__main__':
    app = ps.Application()
    doc = app.activeDocument

    lines_layers = {
        21: doc.artLayers.getByName(f'__ФИО21'),
        22: doc.artLayers.getByName(f'__ФИО22'),
        31: doc.artLayers.getByName(f'__ФИО31'),
        32: doc.artLayers.getByName(f'__ФИО32'),
        33: doc.artLayers.getByName(f'__ФИО33'),
        41: doc.artLayers.getByName(f'__ФИО41'),
        42: doc.artLayers.getByName(f'__ФИО42'),
        43: doc.artLayers.getByName(f'__ФИО43'),
        44: doc.artLayers.getByName(f'__ФИО44'),
    }

    BASE_PATH = Path(__file__).parent

    parser = argparse.ArgumentParser()
    parser.add_argument('--path', type=Path, nargs='?', default=BASE_PATH / 'studs.xlsx')
    parser.add_argument('--results-dir', type=Path, nargs='?', default=BASE_PATH / 'results')
    parser.add_argument('--lines', type=int, nargs='*', default=[1, 2, 3, 4])
    parser.add_argument('--id-mode', action='store_true', default=False)
    options = parser.parse_args().__dict__
    rd = options['results_dir']
    if not rd.exists():
        rd.mkdir()
    assert rd.is_dir()
    main(**options)
